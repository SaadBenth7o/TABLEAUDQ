"""
excel_parser.py
───────────────
Reads the DQ checks from the Excel worksheet (Sheet1: Quality_checks_poc) and
produces a structured `checks_config.yaml` file that is used by every subsequent
run — so the Excel file is only touched once (or when explicitly refreshed with
the --refresh-config flag in run_dq.py).

Responsibilities:
  1. Parse each row into a typed `DQCheck` dataclass.
  2. Extract the Virtualisation *table name* from the full column path.
     e.g. "VIRTUALISATION.staging-nova-referentieltiers.professional_description.turnover"
          → table: "professional_description", column: "turnover"
  3. Group checks by Virtualisation table name.
  4. Deduplicate SQL queries: identical SQL strings are executed only once per
     run; the result is then mapped back to every check that shares that SQL.
  5. Serialise the resulting structure to `checks_config.yaml`.
"""

from __future__ import annotations

import hashlib
import logging
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

import yaml
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

logger = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────

EXCEL_FILE: str = os.getenv("EXCEL_FILE", "Stewardship_Workshop_Template_Tiers_Counterparties.xlsx")
EXCEL_SHEET: str = os.getenv("EXCEL_SHEET", "Quality_checks_poc")
CONFIG_OUTPUT: Path = Path("checks_config.yaml")

# Column indices in the sheet (0-based)
COL_DOMAIN = 0       # Code Domaine (Triplet)
COL_DATASET = 1      # Nom du Dataset/Table
COL_DREMIO_COL = 2   # Champs/Colonne (Dremio)
COL_VIRT_COL = 3     # Champs/Colonne (Virtualisation)  ← full path
COL_RAW_COL = 4      # Champs/Colonne (Raw ou BDD SOP)
COL_RULE = 5         # Règle Data Quality
COL_SQL = 6          # SQL query for the rule


# ── Data model ────────────────────────────────────────────────────────────────

@dataclass
class DQCheck:
    """Represents a single data quality check extracted from one Excel row."""

    domain: str                  # e.g. "PTP-TIE-POR"
    dataset: str                 # e.g. "CIHOne.CLIENTS.Professionnels.clients"
    dremio_col: str              # field name in Dremio / CIHOne
    virt_table: str              # Virtualisation table name (extracted)
    virt_col: str                # Virtualisation column name (extracted)
    virt_full_path: str          # full Virtualisation path as written in Excel
    raw_col: str                 # field name in raw / SOP database
    rule: str                    # e.g. "Completude"
    sql: str                     # SQL query string
    query_id: str = field(init=False)  # SHA-256 fingerprint of the SQL (first 12 chars)

    def __post_init__(self) -> None:
        # Stable short ID based on SQL content — used to deduplicate queries
        self.query_id = hashlib.sha256(self.sql.strip().encode()).hexdigest()[:12]


@dataclass
class UniqueQuery:
    """One deduplicated SQL query and the list of checks that reference it."""

    query_id: str
    sql: str
    used_by_checks: List[str] = field(default_factory=list)  # list of dremio_col names


# ── Parsing helpers ────────────────────────────────────────────────────────────

def _extract_virt_table_and_col(full_path: Optional[str]) -> tuple[str, str]:
    """
    Extract the Virtualisation table name and column name from the full path.

    Examples:
      "VIRTUALISATION.staging-nova-referentieltiers.professional_description.turnover"
        → ("professional_description", "turnover")
      "VIRTUALISATION.staging-nova-referentieltiers.customers.customer_type"
        → ("customers", "customer_type")

    Falls back to ("unknown", full_path) if the path cannot be parsed.
    """
    if not full_path:
        return "unknown", ""

    parts = str(full_path).strip().split(".")
    # Expected structure: VIRTUALISATION . <schema> . <table> . <column>
    #                      idx 0            idx 1      idx 2     idx 3
    if len(parts) >= 4:
        return parts[-2], parts[-1]
    if len(parts) == 3:
        return parts[-1], ""
    return "unknown", full_path


def _clean(value) -> str:
    """Return a stripped string, or empty string for None/whitespace values."""
    return str(value).strip() if value is not None else ""


# Map of accented rule names coming from the Excel to their normalised equivalents.
# This lets us keep the Excel file unchanged while ensuring outputs are accent-free.
_RULE_NORMALISATION: dict = {
    "Completude": "Complétude",
}


def _normalize_rule(rule: str) -> str:
    """Normalise a rule name read from Excel (strip accents as per _RULE_NORMALISATION)."""
    return _RULE_NORMALISATION.get(rule, rule)


# ── Main parser ───────────────────────────────────────────────────────────────

def parse_excel(excel_path: Optional[str] = None, sheet_name: Optional[str] = None) -> List[DQCheck]:
    """
    Open the Excel workbook and parse every data row (skipping the header) into
    a list of `DQCheck` objects.  Rows with an empty SQL column are skipped with
    a warning.

    Args:
        excel_path: Path to the .xlsx file.  Defaults to EXCEL_FILE env var.
        sheet_name: Worksheet name.  Defaults to EXCEL_SHEET env var.

    Returns:
        List of DQCheck objects, one per valid row.
    """
    path = excel_path or EXCEL_FILE
    sheet = sheet_name or EXCEL_SHEET

    logger.info("Opening workbook: %s  →  sheet: %s", path, sheet)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]

    checks: List[DQCheck] = []
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError(f"Sheet '{sheet}' is empty.")

    skipped_non_virt = 0

    # Skip header row (index 0)
    for row_idx, row in enumerate(rows[1:], start=2):
        sql = _clean(row[COL_SQL] if len(row) > COL_SQL else None)
        if not sql:
            logger.warning("Row %d skipped — no SQL query found.", row_idx)
            continue

        # ── VIRTUALISATION filter ──────────────────────────────────────────────
        # Only keep checks whose SQL queries the VIRTUALISATION layer.
        # Rows whose SQL targets CIHOne or any other source directly are excluded.
        if "VIRTUALISATION" not in sql.upper():
            skipped_non_virt += 1
            logger.debug(
                "Row %d skipped — SQL does not target VIRTUALISATION (sql preview: %.80s…)",
                row_idx,
                sql.replace("\n", " "),
            )
            continue

        virt_full = _clean(row[COL_VIRT_COL] if len(row) > COL_VIRT_COL else None)
        virt_table, virt_col = _extract_virt_table_and_col(virt_full)

        check = DQCheck(
            domain=_clean(row[COL_DOMAIN] if len(row) > COL_DOMAIN else None),
            dataset=_clean(row[COL_DATASET] if len(row) > COL_DATASET else None),
            dremio_col=_clean(row[COL_DREMIO_COL] if len(row) > COL_DREMIO_COL else None),
            virt_table=virt_table,
            virt_col=virt_col,
            virt_full_path=virt_full,
            raw_col=_clean(row[COL_RAW_COL] if len(row) > COL_RAW_COL else None),
            rule=_normalize_rule(_clean(row[COL_RULE] if len(row) > COL_RULE else None)),
            sql=sql,
        )
        checks.append(check)

    if skipped_non_virt:
        logger.info(
            "Skipped %d row(s) whose SQL did not target VIRTUALISATION.",
            skipped_non_virt,
        )

    wb.close()
    logger.info(
        "Parsed %d VIRTUALISATION checks from %d data rows (%d non-VIRTUALISATION skipped).",
        len(checks),
        len(rows) - 1,
        skipped_non_virt,
    )
    return checks


def build_unique_queries(checks: List[DQCheck]) -> Dict[str, UniqueQuery]:
    """
    Deduplicate SQL queries across all checks.

    Returns:
        Dict mapping query_id → UniqueQuery (SQL + list of check labels that use it).
    """
    unique: Dict[str, UniqueQuery] = {}
    for check in checks:
        if check.query_id not in unique:
            unique[check.query_id] = UniqueQuery(query_id=check.query_id, sql=check.sql)
        unique[check.query_id].used_by_checks.append(
            f"{check.virt_table}.{check.virt_col}"
        )
    logger.info(
        "Deduplicated %d checks → %d unique SQL queries.", len(checks), len(unique)
    )
    return unique


# ── Config serialiser ─────────────────────────────────────────────────────────

def _checks_to_dict(checks: List[DQCheck]) -> dict:
    """
    Build the full config dictionary that will be written to checks_config.yaml.

    Structure:
      generated_at: <ISO timestamp>
      source_file:  <excel path>
      source_sheet: <sheet name>
      tables:
        <table_name>:
          checks:
            - domain, dataset, dremio_col, virt_col, raw_col, rule, sql, query_id
      unique_queries:
        <query_id>:
          sql: ...
          used_by: [...]
    """
    from datetime import datetime

    tables: Dict[str, list] = {}
    for c in checks:
        tables.setdefault(c.virt_table, []).append({
            "domain": c.domain,
            "dataset": c.dataset,
            "dremio_col": c.dremio_col,
            "virt_col": c.virt_col,
            "virt_full_path": c.virt_full_path,
            "raw_col": c.raw_col,
            "rule": c.rule,
            "sql": c.sql,
            "query_id": c.query_id,
        })

    unique_queries = build_unique_queries(checks)

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_file": EXCEL_FILE,
        "source_sheet": EXCEL_SHEET,
        "tables": {
            tbl: {"checks": rows} for tbl, rows in tables.items()
        },
        "unique_queries": {
            qid: {"sql": uq.sql, "used_by": uq.used_by_checks}
            for qid, uq in unique_queries.items()
        },
    }


def generate_config(
    excel_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
    output_path: Optional[Path] = None,
) -> Path:
    """
    Parse the Excel file and write `checks_config.yaml`.

    Args:
        excel_path:  Path to the .xlsx file (defaults to env EXCEL_FILE).
        sheet_name:  Worksheet name (defaults to env EXCEL_SHEET).
        output_path: Where to write the YAML (defaults to `checks_config.yaml`).

    Returns:
        Path to the generated config file.
    """
    checks = parse_excel(excel_path, sheet_name)
    config_dict = _checks_to_dict(checks)

    dest = output_path or CONFIG_OUTPUT
    with open(dest, "w", encoding="utf-8") as fh:
        yaml.dump(config_dict, fh, allow_unicode=True, sort_keys=False, default_flow_style=False)

    logger.info("Config written to: %s", dest)
    return dest


def load_config(config_path: Optional[Path] = None) -> dict:
    """
    Load an existing `checks_config.yaml` from disk.

    Args:
        config_path: Path to the YAML file (defaults to `checks_config.yaml`).

    Returns:
        Parsed config dictionary.

    Raises:
        FileNotFoundError: If the config file does not exist yet (run with --refresh-config).
    """
    path = config_path or CONFIG_OUTPUT
    if not path.exists():
        raise FileNotFoundError(
            f"Config file not found: {path}\n"
            "Run `python lib/run_dq.py --refresh-config` to generate it from the Excel file."
        )
    with open(path, encoding="utf-8") as fh:
        return yaml.safe_load(fh)


# ── CLI entry point (for standalone testing) ──────────────────────────────────

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
    dest = generate_config()
    print(f"\n✓ checks_config.yaml generated → {dest}")
